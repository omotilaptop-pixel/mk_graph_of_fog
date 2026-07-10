# -*- coding: utf-8 -*-
"""
気象庁ダウンロード形式データ（只見_2026年1_5月_気象データ.xlsx など）を
自動で読み込んでグラフ化するプログラム。

【読み込む気象データ】
  B列: 気温(℃)
  C列: 降水量(mm)
  D列: 風速(m/s)
  G列: 相対湿度(％)
  H列: 露点温度(℃)

【J〜AP列（現象コード）】
  "/"      = その時刻に現象なし
  空白     = まだデータ未入力
  1〜10    = 現象発生コード（下記の意味）
      1  薄い川霧
      2  川霧
      3  濃い川霧
      4  薄い全体霧
      5  全体霧
      6  全体濃い霧
      7  薄い層雲
      8  濃い層雲
      9  霧雨
      10 雨

【出力グラフ（5項目 × 月ごとに分割）】
  ① 気温 × 現象コード（「/」・1〜10）
  ② 降水量 × 現象コード
  ③ 風速 × 現象コード
  ④ 相対湿度 × 現象コード
  ⑤ 露点温度 × 現象コード

  各グラフは、気象データの時系列（線 or 棒グラフ）と、
  その時刻に記録された現象コード（「/」・1〜10）を“レーン”形式で
  分けて表示する2段構成です。データ期間が長い場合に1枚の画像へ
  詰め込みすぎて見づらくならないよう、「年-月」ごとにファイルを
  分割して出力します（例: 5ヶ月分のデータなら、1項目につき5枚、
  合計25枚のPNGが生成されます）。

使い方:
    python3 weather_visualizer.py 入力ファイル.xlsx [出力フォルダ]

Jupyter / Google Colab で直接セルに貼って実行する場合は、
下の DEFAULT_INPUT_FILE / DEFAULT_OUTPUT_DIR を編集してください。
（ファイルが見つからない場合、Colabなら自動でアップロード画面が出ます）
"""
#　=======================================================================
#　グラフに関しての説明
#　=======================================================================

"""
生成されたグラフでは、結合された気象要素（気温、相対湿度など）と日付、気象コード（霧発生など）すべてが適応しているかどうかの整合性チェックはしました。
グラフの見方は、日付が書いてあるメモリの位置はその日の正午となっています。
使用できるファイスの形式は、teamsの投稿にあがっているような形式のみです。
現象コードの色一覧

　＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝
　 /　薄い黄色　「現象なし」
 　1　薄い水色　「薄い川霧」
 　2　水色　　　「川霧」
 　3　濃い水色　「濃い川霧」
 　4　薄い灰色　「薄い全体霧」
 　5　灰色　　　「全体霧」
 　6　紺色　　　「全体濃い霧」
 　7　オレンジ色「薄い層雲」
 　8　赤色　　　「濃い層雲」
 　9　薄い緑色　「霧雨」
 　10 緑色　　　「雨」
　＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝
"""

import sys
import os
import re
import glob
import subprocess
import datetime
import time

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")  # 画面のないサーバー環境（Colab等）でも背景で安全にグラフを描画するための設定
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.font_manager as fm
from matplotlib.patches import Patch
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


# ===========================================================================
# 0. 日本語フォント自動設定・インストール機能（文字化け・豆腐文字対策）
# ===========================================================================

# Linux/Google Colab環境等で標準的な日本語フォントのキーワードリスト
_CJK_FONT_KEYWORDS = [
    "noto sans cjk", "noto serif cjk", "ipaex", "ipagothic", "ipa gothic",
    "takao", "vl gothic", "yu gothic", "ms gothic", "hiragino",
    "source han sans", "droid sans fallback",
]

def _find_cjk_font():
    """システム内に既にインストールされている日本語フォントを探して名前を返す"""
    for f in fm.fontManager.ttflist:
        if any(k in f.name.lower() for k in _CJK_FONT_KEYWORDS):
            return f.name
    return None

def _install_noto_cjk_font():
    """Linux環境（Google Colabなど）の場合に、コマンド経由で日本語フォントを自動インストールする"""
    try:
        # aptパッケージマネージャーの更新
        subprocess.run(["apt-get", "update"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=120, check=False)
        # Noto Sans CJK 日本語フォントをインストール
        result = subprocess.run(
            ["apt-get", "install", "-y", "fonts-noto-cjk"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=300, check=False,
        )
        return result.returncode == 0
    except Exception:
        return False

def setup_japanese_font():
    """matplotlibの日本語設定をコントロールするメイン関数"""
    font_name = _find_cjk_font()

    # 日本語フォントがシステムにない場合は自動でダウンロード・インストールを試みる
    if font_name is None:
        print("日本語フォントが見つからないため、自動インストールを試みます…")
        time.sleep(1)
        print(" この処理には時間がかかることがあります")
        if _install_noto_cjk_font():
            # インストールしたフォントファイルをmatplotlibに再認識させる
            for fp in glob.glob("/usr/share/fonts/**/*.[ot]t[fc]", recursive=True):
                try:
                    fm.fontManager.addfont(fp)
                except Exception:
                    pass
            font_name = _find_cjk_font()

    # フォントの設定適用
    if font_name:
        plt.rcParams["font.family"] = font_name
        print(f"日本語フォントを設定しました: {font_name}")
    else:
        print("【警告】日本語フォントの自動設定に失敗しました。グラフ内の文字が『□』になる可能性があります。")

    # グラフ内でマイナス記号「-」が文字化けするのを防ぐ設定
    plt.rcParams["axes.unicode_minus"] = False
    return font_name

# スクリプト起動時にフォント設定を即座に実行
setup_japanese_font()


# ===========================================================================
# 1. Excelの列レイアウト設定と現象コード（色・ラベル）の定義
# ===========================================================================

# メインとなる気象要素（只見）が格納されているExcelの列名（アルファベット）
COL_DATETIME = "A"    # 年月日時
COL_TEMP = "B"        # 気温(℃)
COL_PRECIP = "E"      # 降水量(mm)
COL_WIND = "H"        # 風速(m/s)
COL_DEWPOINT = "S"    # 露点温度(℃)
COL_HUMID = "V"       # 相対湿度(％)

# 各観測地点ごとの「現象コード」が並んでいる列の開始と終了の範囲
PHENOMENA_RANGE = ("AC", "BH")  # AC列からBH列まで（計32地点分）

# プログラム内部で処理しやすいように列名とグラフ用のラベルをマッピング
MAIN_COLUMNS = {
    COL_TEMP: "気温(℃)",
    COL_PRECIP: "降水量(mm)",
    COL_WIND: "風速(m/s)",
    COL_DEWPOINT: "露点温度(℃)",
    COL_HUMID: "相対湿度(％)",
}

# 現象コード（1〜10）に対応する正式な「現象名」
PHENOM_LABELS = {
    1: "薄い川霧", 2: "川霧", 3: "濃い川霧", 4: "薄い全体霧", 5: "全体霧",
    6: "全体濃い霧", 7: "薄い層雲", 8: "濃い層雲", 9: "霧雨", 10: "雨",
}

# 現象コード（1〜10）をグラフに描画する際の見やすい配色（カラーコード）
PHENOM_COLORS = {
    1: "#aed6f1", 2: "#3498db", 3: "#1a5276", 4: "#dcdde1", 5: "#909497",
    6: "#2c3e50", 7: "#f8c471", 8: "#d35400", 9: "#a9dfbf", 10: "#196f3d",
}

# 現象コードが「/」（現象なし）だった時間帯の背景色（薄い黄色）
SLASH_COLOR = "#f9e79f"


# ===========================================================================
# 2. データ読み込みおよびデータクレンジング（品質情報等の除外）
# ===========================================================================

def find_header_row(ws, search_col="A", keyword="年月日時", max_search_rows=15):
    """Excelの先頭数行から、『年月日時』というヘッダーが書かれた実際の行番号を自動で探す"""
    col_idx = column_index_from_string(search_col)
    for r in range(1, max_search_rows + 1):
        v = ws.cell(row=r, column=col_idx).value
        if v == keyword:
            return r
    raise ValueError(f"ヘッダー行（{search_col}列に「{keyword}」）が見つかりませんでした。気象庁のデータか確認してください。")


def load_weather_data(filepath, sheet_name=None):
    """Excelファイルを解析し、メインデータ、各地点の現象データ、地点名のリストを作成して返す"""
    # Excelファイルを読み込み（計算式ではなく値そのものを読み込む設定）
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    # ヘッダー行（年月日時）の位置と、データが始まる行（その一つき下）を決定
    header_row = find_header_row(ws, COL_DATETIME)
    data_start_row = header_row + 1
    dt_col_idx = column_index_from_string(COL_DATETIME)

    # 現象コードの対象列（AC列〜BH列）の範囲をループ処理用にアルファベットの配列へ変換
    start_idx = column_index_from_string(PHENOMENA_RANGE[0])
    end_idx = column_index_from_string(PHENOMENA_RANGE[1])
    phenom_cols = [get_column_letter(c) for c in range(start_idx, end_idx + 1)]

    # 【重要】Excelの3行目から、AC〜BH列にそれぞれ対応する「観測地点名」を自動抽出して記憶
    location_mapping = {}
    for col_letter in phenom_cols:
        cidx = column_index_from_string(col_letter)
        loc_name = ws.cell(row=3, column=cidx).value
        if loc_name:
            # 改行コードや前後の不要な空白を除去して綺麗な文字列にする
            loc_name = str(loc_name).strip().replace("\n", "").replace("\r", "")
        else:
            loc_name = f"地点_{col_letter}"
        location_mapping[col_letter] = loc_name

    main_rows = []
    phenom_rows = []

    # データのメインループ：行単位でExcelをスキャン
    for r in range(data_start_row, ws.max_row + 1):
        date_val = ws.cell(row=r, column=dt_col_idx).value
        if date_val is None:
            continue  # 日時が空の行は無視

        # ① 只見の主要な気象要素（気温・湿度など）を抽出して配列に格納
        mrow = {"datetime": date_val}
        for col_letter, label in MAIN_COLUMNS.items():
            cidx = column_index_from_string(col_letter)
            mrow[label] = ws.cell(row=r, column=cidx).value
        main_rows.append(mrow)

        # ② 各地点の現象コードを抽出して配列に格納
        prow = {"datetime": date_val}
        for col_letter in phenom_cols:
            cidx = column_index_from_string(col_letter)
            prow[col_letter] = ws.cell(row=r, column=cidx).value
        phenom_rows.append(prow)

    # ――― データ構造の整理とデータクレンジング ―――

    # メインの気象データをPandasのデータフレーム形式に集計
    main_df = pd.DataFrame(main_rows)
    # A列のデータを日付・時刻型に一括変換（変換できない文字はエラーではなく空（NaT）にする）
    main_df["datetime"] = pd.to_datetime(main_df["datetime"], errors="coerce")
    # 【対策】「品質情報」や「均質番号」など、日付に変換できない不要な文字列行を確実に除外
    main_df = main_df.dropna(subset=["datetime"])
    # 気温や降水量などの値を数値型に一括変換（「/」などの文字列は自動的に欠測値 NaN に置換される）
    for label in MAIN_COLUMNS.values():
        main_df[label] = pd.to_numeric(main_df[label], errors="coerce")
    # 時系列順（古い順）に確実に並び替える
    main_df = main_df.sort_values("datetime").reset_index(drop=True)

    # 現象データ側も同様にクレンジング処理を実行
    phenom_df = pd.DataFrame(phenom_rows)
    phenom_df["datetime"] = pd.to_datetime(phenom_df["datetime"], errors="coerce")
    phenom_df = phenom_df.dropna(subset=["datetime"])
    phenom_df = phenom_df.sort_values("datetime").reset_index(drop=True)

    return main_df, phenom_df, phenom_cols, location_mapping


def encode_phenomena_cell(value):
    """Excelの現象セルに入っている複雑な文字（数字やスラッシュ、空欄）をきれいな数値コードに変換する"""
    if value is None:
        return np.nan
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if s == "/":
        return 0.0  # 現象なし「/」はプログラム内部で 0.0 として扱う

    # 文字列の中から数字の部分だけを抜き出す（例: "1 (品質4)" のようなケースから "1" を抽出）
    nums = re.findall(r"\d+", s)
    if nums:
        return float(max(int(n) for n in nums))  # 含まれる最大の数値を現象コードとする
    return np.nan


# ===========================================================================
# 3. グラフ生成・描画処理（上下2段構成・タイムライン同期システム）
# ===========================================================================

# グラフの横幅を自動計算するための基準値（1インチあたり何日分を表示させるか）
DAYS_PER_INCH = 0.55
MIN_FIG_WIDTH = 16   # グラフの最小横幅（インチ）
MAX_FIG_WIDTH = 60   # グラフの最大横幅（インチ）
SAVE_DPI = 150       # 画像として保存する際の解像度（画質）


def compute_fig_width(times):
    """データの日数に応じて、グラフが横に潰れない最適な横幅（サイズ）を自動計算する"""
    total_days = (times.max() - times.min()).total_seconds() / 86400.0
    width = max(total_days, 1.0) / DAYS_PER_INCH
    return float(np.clip(width, MIN_FIG_WIDTH, MAX_FIG_WIDTH))


def split_by_month_two(main_df, phenom_df):
    """複数月（1月〜5月など）の塊のデータを、月ごと（1ヶ月単位）のデータに自動でハサミで切り分ける"""
    m = main_df.copy()
    p = phenom_df.copy()
    m["__ym"] = m["datetime"].dt.strftime("%Y-%m")
    p["__ym"] = p["datetime"].dt.strftime("%Y-%m")
    groups = []
    # 存在するユニークな「年-月」ごとにデータを抽出してリスト化
    for ym in sorted(set(m["__ym"]) | set(p["__ym"])):
        msub = m[m["__ym"] == ym].drop(columns="__ym").reset_index(drop=True)
        psub = p[p["__ym"] == ym].drop(columns="__ym").reset_index(drop=True)
        groups.append((ym, msub, psub))
    return groups


def _draw_lane_panel(ax2, phenom_df, phenom_cols, location_mapping):
    """下段パネルに全観測地点（32地点）を縦に並べて、現象コードの発生履歴をレーン形式で描画する"""
    ptimes = phenom_df["datetime"]
    n_lanes = len(phenom_cols)

    # 1地点（1列）ずつ、下から上に向かって縦に並べて描画していく
    for row, col_letter in enumerate(phenom_cols):
        # 該当列のデータを数値コード（0〜10）に変換して取得
        col_values = phenom_df[col_letter].map(encode_phenomena_cell).to_numpy()

        # ーーー 【重要】時間軸のバグ修正処理 ーーー
        # matplotlibのvlines（縦線を描く関数）は、日付オブジェクトのままだと横軸の位置がズレる特性があります。
        # そのため、mdates.date2num()を使い、日付を「正確な数値の座標」に明示的に変換して渡しています。

        # 現象なし「/」の描画（薄い黄色の背景線を引く）
        slash_mask = col_values == 0
        if slash_mask.any():
            x_vals = mdates.date2num(ptimes[slash_mask])
            ax2.vlines(x_vals, row - 0.35, row + 0.35,
                       color=SLASH_COLOR, linewidth=1.0, alpha=0.9, zorder=2)

        # 現象コード 1〜10 の描画（それぞれの定義色で太い縦線を引く）
        for code in range(1, 11):
            code_mask = col_values == code
            if code_mask.any():
                x_vals = mdates.date2num(ptimes[code_mask])
                ax2.vlines(x_vals, row - 0.42, row + 0.42,
                           color=PHENOM_COLORS[code], linewidth=2.4, zorder=3)

    # 地点ごとの境界線（うすい横線）を引いて見やすくする
    for row in range(n_lanes + 1):
        ax2.axhline(row - 0.5, color="#ececec", linewidth=0.5, zorder=1)

    # 縦軸（Y軸）の目盛りに、Excelの3行目から取得した実際の「地点名」をセット
    ax2.set_yticks(range(n_lanes))
    ax2.set_yticklabels([location_mapping[c] for c in phenom_cols], fontsize=8)
    # 上が1番目の地点、下が最後の地点になるように軸の向きを反転固定する
    ax2.set_ylim(n_lanes - 0.5, -0.5)
    ax2.set_xlabel("日時")
    ax2.set_ylabel("現象記録地点", fontsize=10)

    # グラフの右側に表示する「現象コードの凡例（カラーチャート）」を作成
    legend_handles = [Patch(facecolor=SLASH_COLOR, label="「/」現象なし")]
    for code in range(1, 11):
        legend_handles.append(Patch(facecolor=PHENOM_COLORS[code], label=f"{code}: {PHENOM_LABELS[code]}"))
    ax2.legend(handles=legend_handles, bbox_to_anchor=(1.0, 1.0), loc="upper left",
               fontsize=8, borderaxespad=0., title="現象コード", title_fontsize=9)


def _finalize_figure(fig, ax1, ax2, times, fig_width):
    """グラフの横軸（日付目盛り）の見た目や、表示範囲の最終調整を行う関数"""
    ax2.xaxis_date()
    # 横軸の表示フォーマットを「月/日」（例: 03/01）にする
    ax2.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d"))

    # グラフの横幅サイズに応じて、日付目盛りが細かくなりすぎたり疎らになりすぎたりしないよう自動調整
    n_ticks_target = max(10, int(fig_width / 1.3))
    ax2.xaxis.set_major_locator(
        mdates.AutoDateLocator(minticks=n_ticks_target, maxticks=n_ticks_target * 2)
    )

    # グラフの左端と右端を、その月の「最初の1時間」から「最後の1時間」へきっちり合わせる（余白の排除）
    ax1.set_xlim(times.min(), times.max())
    # 斜め文字にして重なりを防ぐ
    fig.autofmt_xdate()


def _make_fig(times, phenom_cols):
    """上下2段構成のグラフの土台（キャンバス）を生成する関数"""
    n_lanes = len(phenom_cols)
    fig_width = compute_fig_width(times)

    # 【体裁調整】32地点という膨大なレーンが縦にきれいに収まり、文字が潰れないよう、地点数に比例して下段の高さを自動拡張する
    lane_panel_height = max(6.0, n_lanes * 0.24)
    fig_height = 5.5 + lane_panel_height

    # 上段と下段のグラフを生成（sharex=True によって上下の時間軸の位置が1分のズレもなく完全に固定同期される）
    fig, (ax1, ax2) = plt.subplots(
        2, 1, figsize=(fig_width, fig_height), sharex=True,
        gridspec_kw={"height_ratios": [5.5, lane_panel_height], "hspace": 0.04}, # 上下の隙間をほぼゼロに詰める
    )
    return fig, ax1, ax2, fig_width


def plot_temp_humid_dew(main_df, phenom_df, phenom_cols, location_mapping, location_name, out_path):
    """【パターン①】気温・露点温度・相対湿度 × 現象コード の組合せグラフを生成する"""
    times = main_df["datetime"]
    fig, ax1, ax2, fig_width = _make_fig(times, phenom_cols)

    # ーーー 上段の描画（気温・露点温度・湿度） ーーー
    # 左軸（℃）に気温と露点温度をプロット
    l1, = ax1.plot(times, main_df["気温(℃)"],    color="#e74c3c", linewidth=1.1, label="気温(℃)", zorder=3)
    l2, = ax1.plot(times, main_df["露点温度(℃)"], color="#16a085", linewidth=1.1, label="露点温度(℃)", zorder=3)
    ax1.set_ylabel("気温・露点温度（℃）", fontsize=10)
    ax1.grid(True, alpha=0.25)

    # 右軸（％）を作成して相対湿度をプロット（目盛り範囲は0〜115%に固定してゆとりを持たせる）
    ax1r = ax1.twinx()
    l3, = ax1r.plot(times, main_df["相対湿度(％)"], color="#8e44ad", linewidth=0.9, alpha=0.65, label="相対湿度(％)", zorder=2)
    ax1r.set_ylabel("相対湿度（％）", fontsize=10)
    ax1r.set_ylim(0, 115)

    # タイトルと凡例の設定
    ax1.set_title(f"【{location_name}】気温・露点温度（左軸℃）・相対湿度（右軸%） と 各地点の現象コードの関係", fontsize=13)
    ax1.legend(handles=[l1, l2, l3], loc="upper left", fontsize=9)

    # ーーー 下段の描画（全地点現象レーン） と 仕上げ ーーー
    _draw_lane_panel(ax2, phenom_df, phenom_cols, location_mapping)
    _finalize_figure(fig, ax1, ax2, times, fig_width)

    # 画像ファイルとしてディスクに書き出し
    fig.savefig(out_path, dpi=SAVE_DPI, bbox_inches="tight")
    plt.close(fig)  # メモリ解放のため明示的にクローズ


def plot_wind_precip(main_df, phenom_df, phenom_cols, location_mapping, location_name, out_path):
    """【パターン②】風速・降水量 × 現象コード の組合せグラフを生成する"""
    times = main_df["datetime"]
    fig, ax1, ax2, fig_width = _make_fig(times, phenom_cols)

    # ーーー 上段の描画（風速・降水量） ーーー
    # 左軸（m/s）に風速を折れ線グラフでプロット
    l1, = ax1.plot(times, main_df["風速(m/s)"], color="#27ae60", linewidth=1.1, label="風速(m/s)", zorder=3)
    ax1.set_ylabel("風速（m/s）", fontsize=10)
    ax1.grid(True, alpha=0.25)

    # 右軸（mm）を作成して降水量を縦棒グラフ（bar）でプロット
    ax1r = ax1.twinx()
    ax1r.bar(times, main_df["降水量(mm)"], width=0.03, color="#2980b9", alpha=0.55, label="降水量(mm)", zorder=2)
    # Y軸の上限値をデータの最大値に応じて適度に自動調整
    pmax = main_df["降水量(mm)"].dropna().max() if not main_df["降水量(mm)"].dropna().empty else 1.0
    ax1r.set_ylim(0, max(pmax * 3.5, 2.0))
    ax1r.set_ylabel("降水量（mm）", fontsize=10)

    # タイトルと凡例の設定
    ax1.set_title(f"【{location_name}】風速（左軸m/s）・降水量（右軸mm） と 各地点の現象コードの関係", fontsize=13)
    ax1.legend(handles=[l1, Patch(color="#2980b9", alpha=0.55, label="降水量(mm)")], loc="upper left", fontsize=9)

    # ーーー 下段の描画（全地点現象レーン） と 仕上げ ーーー
    _draw_lane_panel(ax2, phenom_df, phenom_cols, location_mapping)
    _finalize_figure(fig, ax1, ax2, times, fig_width)

    # 画像ファイルとしてディスクに書き出し
    fig.savefig(out_path, dpi=SAVE_DPI, bbox_inches="tight")
    plt.close(fig)


def plot_combo_by_month(main_df, phenom_df, phenom_cols, location_mapping, location_name, out_dir):
    """読み込んだ全データを月ごとに切り分け、2つのパターンのグラフをそれぞれ連続出力する制御関数"""
    for ym, msub, psub in split_by_month_two(main_df, phenom_df):
        # その月のデータが1件以下（データがないなど）の場合はスキップ
        if msub["datetime"].nunique() < 2:
            continue

        # グラフ内に表示する地点名と対象年月のラベル（例: 只見（2026-03））
        loc = f"{location_name}（{ym}）"

        # ① 気温・湿度・露点温度のグラフを生成
        plot_temp_humid_dew(
            msub, psub, phenom_cols, location_mapping, loc,
            os.path.join(out_dir, f"{location_name}_①気温・湿度・露点×現象コード_{ym}.png"),
        )
        # ② 風速・降水量のグラフを生成
        plot_wind_precip(
            msub, psub, phenom_cols, location_mapping, loc,
            os.path.join(out_dir, f"{location_name}_②風速・降水量×現象コード_{ym}.png"),
        )


# ===========================================================================
# 4. メイン実行部および Google Colab 支援機能
# ===========================================================================

# 標準のファイル名と出力先フォルダの設定（ファイルが見つからない場合のデフォルト挙動用）
DEFAULT_INPUT_FILE = "気象データ.xlsx"
DEFAULT_OUTPUT_DIR = "./output_graphs"


def _try_colab_upload():
    """Google Colab特有のUIを起動し、ブラウザからExcelファイルを直接アップロードさせる関数"""
    if "google.colab" not in sys.modules:
        return None
    try:
        from google.colab import files as colab_files
    except ImportError:
        return None

    print("【確認】入力ファイルが指定のパスに見つかりません。")
    time.sleep(1)
    print("アップロード画面を表示しますので、解析したい気象データのExcelファイルを選択してください…")
    uploaded = colab_files.upload()
    for name in uploaded.keys():
        if name.lower().endswith((".xlsx", ".xls")):
            return os.path.abspath(name)  # アップロードされたファイルの絶対パスを返す
    return None


def _resolve_input_file(filepath):
    """指定されたパスにファイルがあるかチェックし、なければColabのアップロードを促す調整関数"""
    if filepath and os.path.isfile(filepath):
        return filepath

    # Google Colab環境の場合はブラウザからの手動アップロード機能を試みる
    if "google.colab" in sys.modules:
        uploaded_path = _try_colab_upload()
        if uploaded_path and os.path.isfile(uploaded_path):
            return uploaded_path

    # ローカルNotebook環境用の対話型パス入力フォールバック
    if "ipykernel" in sys.modules or "IPython" in sys.modules:
        try:
            entered = input("入力ファイルのパスを入力してください（Enterでキャンセル): ").strip()
            if entered and os.path.isfile(entered):
                return entered
        except Exception:
            pass
    return None


def _parse_args():
    """コマンドライン引数（またはスクリプト実行時引数）からファイルパスを自動判定する"""
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    args = [a for a in args if a.lower().endswith((".xlsx", ".xls")) or os.path.isdir(a)]

    filepath = None
    out_dir = None
    for a in args:
        if a.lower().endswith((".xlsx", ".xls")): filepath = a
        else: out_dir = a

    if filepath is None: filepath = DEFAULT_INPUT_FILE
    if out_dir is None: out_dir = DEFAULT_OUTPUT_DIR
    return filepath, out_dir


def main():
    """プログラム全体の実行をコントロールする最上位のエントリーポイント"""
    # 実行引数とファイルの存在解決
    filepath, out_dir = _parse_args()
    filepath = _resolve_input_file(filepath)

    if filepath is None:
        print("【エラー】入力ファイルが指定されなかったか、見つかりませんでした。")
        print("ファイル名が正しく『気象データ.xlsx』になっているか、またはフォルダーに正しく配置されているか確認してください。")
        sys.exit(1)

    # 出力先フォルダ（./output_graphs）が存在しない場合は自動作成
    os.makedirs(out_dir, exist_ok=True)

    print(f"Excelファイルを確認中: {filepath}")
    # データのロードとクレンジング処理を呼び出し
    main_df, phenom_df, phenom_cols, location_mapping = load_weather_data(filepath)

    # ファイル名からコアとなる地点名を自動判定（例: 只見_2026年...xlsx -> 只見）
    base = os.path.splitext(os.path.basename(filepath))[0]
    location_name = base.split("_")[0].split(" ")[0] if ("_" in base or " " in base) else base

    print(f"データ解析成功: {location_name} (計 {len(phenom_cols)} 地点の現象レーンを検出しました)")
    print("月ごと・全地点統合グラフの自動生成を開始します...")
    time.sleep(1)
    print("このグラフの生成には3分ほどの時間がかかることがあります")

    # メインの描画ループ関数を実行
    plot_combo_by_month(main_df, phenom_df, phenom_cols, location_mapping, location_name, out_dir)

    print(f"\nすべてのグラフが正常に生成されました！")
    print(f"出力先フォルダ: {out_dir} の中を確認してください。")


if __name__ == "__main__":
    # このスクリプトが直接実行された場合にmain関数を動かす
    main()

